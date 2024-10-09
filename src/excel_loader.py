import hashlib
import warnings
from typing import Any, Dict, List, Optional, Tuple, Type

import pandas as pd
import openpyxl
from styleframe import StyleFrame

from excel_game import (
    ExcelGame,
    ExcelGameBuilder,
    ExcelRegion as Region,
    FuzzyDateType,
    Playability,
    PlayingStatus,
    TranslationStatus,
)


class ExcelLoader:
    _sf: Optional[StyleFrame] = None
    _cg_df: Optional[pd.DataFrame] = None
    _goo_df: Optional[pd.DataFrame] = None
    _games: Optional[List[ExcelGame]] = None
    _completed_games: Optional[List[ExcelGame]] = None
    _games_on_order: Optional[List[ExcelGame]] = None

    __sheet_path: str

    __BASE_DROPBOX_FOLDER = "C:\\Users\\zachd\\Dropbox\\Video Game Lists"
    __EXCEL_SHEET_NAME = "Games Master List - Final.xlsx"

    def __init__(self, file_path: Optional[str] = None, eager: bool = False):
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        if file_path is None:
            file_path = f"{self.__BASE_DROPBOX_FOLDER}\\{self.__EXCEL_SHEET_NAME}"

        self.__sheet_path = file_path

        if eager:
            self.__load_games()

    def __read_excel(self, file_path: str, sheet: str = "Games") -> StyleFrame:
        return StyleFrame.read_excel(
            file_path,
            sheet_name=sheet,
            keep_default_na=False,
            read_style=True,
            use_openpyxl_styles=False,
        )

    @property
    def df(self) -> pd.DataFrame:
        return self.sf.data_df

    @property
    def sf(self) -> StyleFrame:
        if self._sf is None:
            self._sf = self.__read_excel(self.__sheet_path)

        return self._sf

    @property
    def cg_df(self) -> pd.DataFrame:
        if self._cg_df is None:
            self._cg_df = pd.read_excel(
                self.__sheet_path,
                sheet_name="Finished Games (Adtl Metadata)",
                keep_default_na=False,
            )

        return self._cg_df

    @property
    def goo_df(self) -> pd.DataFrame:
        if self._goo_df is None:
            self._goo_df = pd.read_excel(
                self.__sheet_path,
                sheet_name="Games On Order",
                keep_default_na=False,
            )

            ws = openpyxl.load_workbook(self.__sheet_path)["Games On Order"]
            links = []
            for i in range(2, ws.max_row + 1):
                # 9th Column == Order #
                cell = ws.cell(row=i, column=9)
                if cell.hyperlink is None:
                    if len(links) < self._goo_df.shape[0]:
                        links.append(None)
                        continue
                    break
                links.append(cell.hyperlink.target)

            self._goo_df["Order Link"] = links

        return self._goo_df

    def __load_games(self, df_override: Optional[pd.DataFrame] = None):
        self._games = []

        df = df_override if df_override is not None else self.df

        for _, row in df.iterrows():
            self._games.append(
                ExcelGameBuilder()
                .with_title(str(row["Title"]))
                .with_platform(str(row["Platform"]))
                .with_release_date(
                    row["Release Date"].value
                    if row["Release Date"] != "Early Access"
                    else None
                )
                .with_release_region(Region(row["Release Region"].value))
                .with_publisher(str(row["Publisher"]))
                .with_developer(str(row["Developer"]))
                .with_franchise(self.__none_or_type(row["Franchise"]))
                .with_genre(str(row["Genre"]))
                .with_vr(self.__none_or_type(row["VR"], bool))
                .with_dlc(self.__none_or_type(row["DLC"], bool))
                .with_translation(
                    TranslationStatus(row["English"])
                    if row["English"] is not None and str(row["English"]).strip() != ""
                    else None
                )
                .with_owned(self.__none_or_type(row["Owned"], bool))
                .with_owned_condition(self.__none_or_type(row["Condition"]))
                .with_date_purchased(
                    row["Date Purchased"].value
                    if str(row["Date Purchased"]).strip() != ""
                    else None
                )
                .with_purchase_price(self.__none_or_type(row["Purchase Price"], float))
                .with_owned_format(self.__none_or_type(row["Format"]))
                .with_completed(bool(row["Completed"]))
                .with_date_completed(
                    row["Date Completed"].value
                    if str(row["Date Completed"]).strip() != ""
                    else None
                )
                .with_completion_time(
                    self.__none_or_type(row["Completion Time"], float)
                )
                .with_rating(self.__none_or_type(row["Rating"], float))
                .with_metacritic_rating(
                    self.__none_or_type(row["Metacritic Rating"], float)
                )
                .with_gamefaqs_rating(
                    self.__none_or_type(row["GameFAQs User Rating"], float)
                )
                .with_notes(self.__none_or_type(row["Notes"], str))
                .with_priority(self.__none_or_type(row["Priority"], int))
                .with_wishlisted(self.__none_or_type(row["Wishlisted"], bool))
                .with_estimated_playtime(
                    self.__none_or_type(row["Estimated Time"], float)
                )
                .with_playing_status(
                    PlayingStatus(row["Playing Status"])
                    if row["Playing Status"] is not None
                    and str(row["Playing Status"]).strip() != ""
                    else None
                )
                .with_playability(Playability(row["Playable"]))
                .with_fuzzy_date(
                    FuzzyDateType.YEAR_ONLY
                    if row["Release Date"].style.bold
                    and row["Release Date"].style.italic
                    else (
                        FuzzyDateType.MONTH_AND_YEAR_ONLY
                        if row["Release Date"].style.italic
                        else None
                    )
                )
                .build()
            )

    @property
    def games(self) -> List[ExcelGame]:
        if self._games is None:
            self.__load_games()

        return self._games

    @property
    def completed_games(self) -> List[ExcelGame]:
        if self._completed_games is None:
            self._completed_games = []

            for _, row in self.cg_df.iterrows():
                self._completed_games.append(
                    ExcelGameBuilder()
                    .with_title(str(row["Game"]))
                    .with_platform(str(row["Platform"]))
                    .with_release_date(row["Release"])
                    .with_release_region(Region(row["Region"]))
                    .with_publisher(str(row["Publisher"]))
                    .with_developer(str(row["Developer"]))
                    .with_franchise(self.__none_or_type(row["Franchise"]))
                    .with_genre(str(row["Genre"]))
                    .with_played_in_vr(self.__none_or_type(row["VR"], bool))
                    .with_collection(self.__none_or_type(row["Collection"]))
                    .with_rating(float(row["Rating"]))
                    .with_completion_number(int(row["#"]))
                    .with_date_completed(
                        row["Date"] if str(row["Date"]).strip() != "" else None
                    )
                    .with_completion_time(self.__none_or_type(row["Play Time"], float))
                    .with_steam_deck(self.__none_or_type(row["Steam Deck"], bool))
                    .with_emulated(self.__none_or_type(row["Emulated"], bool))
                    .with_metacritic_rating(
                        self.__none_or_type(row["Critic Score"], float)
                    )
                    .with_completion_notes(self.__none_or_type(row["Notes"]))
                    .build()
                )

        return self._completed_games

    @property
    def games_on_order(self) -> List[ExcelGame]:
        if self._games_on_order is None:
            self._games_on_order = []

            for _, row in self.goo_df.iterrows():
                self._games_on_order.append(
                    ExcelGameBuilder()
                    .with_title(str(row["Title"]))
                    .with_platform(
                        str(row["Platform"]) if row["Platform"] != "TBD" else None
                    )
                    .with_order_vendor(str(row["Vendor"]))
                    .with_date_purchased(row["Ordered Date"])
                    .with_purchase_price(float(row["Price"]))
                    .with_owned_format(str(row["Format"]))
                    .with_order_status(str(row["Status"]))
                    .with_estimated_release(
                        row["Estimated Release"]
                        if row["Estimated Release"] is not None
                        and row["Estimated Release"] != "N/A"
                        else None
                    )
                    .with_order_id(str(row["Order #"]))
                    .with_order_link(str(row["Order Link"]))
                    .with_address_on_order(
                        str(row["Address on Order"])
                        if row["Address on Order"] != "N/A"
                        else None
                    )
                    .with_tracking_number(self.__none_or_type(row["Tracking #"]))
                    .build()
                )

        return self._games_on_order

    @staticmethod
    def __none_or_type(value: Optional[Any], _type: Type = str):
        return None if value is None or str(value).strip() == "" else _type(value)

    def override_sheet(self, df_override: pd.DataFrame):
        self.__load_games(df_override)

    def clear(self):
        self._sf = None
        self._cg_df = None
        self._goo_df = None
        self._games = None
        self._completed_games = None
        self._games_on_order = None

    def reload(self):
        self.clear()
        _ = self.games
        _ = self.completed_games
        _ = self.games_on_order

    def merge(self) -> Tuple[List[ExcelGame], List[Tuple[ExcelGame, str]]]:
        hashed_games: Dict[int, List[ExcelGame]] = {}
        collection_games: Dict[int, List[ExcelGame]] = {}
        collection_hashes: Dict[str, int] = {}
        errors: List[Tuple[ExcelGame, str]] = []

        def add_children(g_hash: int, g: ExcelGame):
            if g_hash in collection_games:
                g.child_games = collection_games[g_hash]

        for g in self.games:
            hashed_g = hash(g)

            if g.collection_hash_id not in collection_hashes:
                collection_hashes[g.collection_hash_id] = hashed_g
            else:
                errors.append((g, "Merged: Duplicate Collection Hash in Main Sheet"))

            if hashed_g in hashed_games:
                hashed_games[hashed_g].append(g)
                errors.append((g, "Merged: Duplicate in Main Sheet"))
            else:
                hashed_games[hashed_g] = [g]

        for g in self.completed_games:
            hashed_g = hash(g)
            if hashed_g in hashed_games:
                hashed_games[hashed_g].append(g)
            else:
                if g.collection is None:
                    errors.append((g, "Merged: Hash Found Only in Completed"))
                elif g.collection_hash_id in collection_hashes:
                    hashed_g = collection_hashes[g.collection_hash_id]
                    if hashed_g in collection_games:
                        collection_games[hashed_g].append(g)
                    else:
                        collection_games[hashed_g] = [g]
                    continue
                else:
                    errors.append((g, "Merged: Collection Hash Not Found"))

        output_games: List[ExcelGame] = []

        for g_hash, games in hashed_games.items():
            if len(games) == 1:
                add_children(g_hash, games[0])
                output_games.append(games[0])
                continue

            if len(games) > 2:
                raise ValueError(
                    f"Unexpectedly found more than 2 games: {', '.join(g.title for g in games)}"
                )

            combined_game = ExcelGame()

            for g in games:
                combined_game = ExcelGameBuilder().merge(g, combined_game)

            add_children(g_hash, combined_game)

            output_games.append(combined_game)

        return (output_games, errors)
